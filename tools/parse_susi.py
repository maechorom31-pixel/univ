#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""수시지원 결과 엑셀 파서 (전남 진학결과 + 나주고 다운로드 원본 2022~2026).

입력 엑셀 파일은 개인정보(실명)를 포함하므로 절대 저장소에 커밋하지 않는다.
이 스크립트의 출력 JSON도 커밋 금지 — 로컬/세션 안에서만 사용한다.
이름은 출력에서 완전히 제거되고 학생은 익명 ID(N2026-001, J-00001 형식)로만 남는다.

사용법:
    python3 tools/parse_susi.py <전남.xlsx> <나주고.xlsx> <출력.json>

출력 스키마는 tools/DESIGN.md 참고.
"""
import json
import re
import sys
import unicodedata

import openpyxl

# ---------------------------------------------------------------- 공통 유틸

FINAL_OUTCOMES = {"합격", "불합격", "충원합격", "합", "불", "충원합"}


def clean(v):
    if v is None:
        return None
    s = unicodedata.normalize("NFC", str(v)).strip()
    return s if s not in ("", "-", "'-", "None") else None


def num(v):
    s = clean(v)
    if s is None:
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def norm_track(v):
    """계열 정규화."""
    s = clean(v)
    if s is None:
        return None
    s = s.replace(" ", "")
    return s if s in ("인문", "자연", "예체능", "공학", "공통") else (s or None)


CAT_MAP = [
    ("학생부위주(교과)", "교과"),
    ("학생부위주(종합)", "종합"),
    ("학생부교과", "교과"),
    ("학생부종합", "종합"),
    ("논술위주", "논술"),
    ("논술", "논술"),
    ("실기/실적위주", "실기"),
    ("실기", "실기"),
    ("수능위주", "수능"),
]


def norm_category(v):
    s = clean(v)
    if s is None:
        return None
    for key, cat in CAT_MAP:
        if key in s:
            return cat
    return "기타"


def norm_final(v):
    """최종단계 → 합격 / 충원합격 / 불합격 / None(미입력)."""
    s = clean(v)
    if s is None:
        return None
    if "충원" in s:
        return "충원합격"
    if s in ("합", "합격"):
        return "합격"
    if s in ("불", "불합격"):
        return "불합격"
    return None


def univ_short(name):
    """'전남대학교(광주)' → '전남대' 식으로 입결 데이터(index.html DATA)의 대학명과 맞춘다.

    입결 쪽 표기 규칙(실측): 대학교→대, 여자대→여대, 교육대→교대, 외국어대→외대,
    과학기술대→과기대, "국립" 접두어는 유지(국립순천대), 분교는 (글)/(세)/(죽전) 등으로 구분.
    """
    s = clean(name)
    if s is None:
        return None
    s = re.sub(r"\s*-\s*.*캠퍼스\s*$", "", s)  # '한국외국어대학교(용인) - 글로벌캠퍼스'
    if re.search(r"\)\s*-", s + "-"):  # '한서대학교(태안) -항공학부' 류: 괄호 뒤 접미 제거
        s = re.sub(r"(\))\s*-\s*[가-힣]+$", r"\1", s)
    campus = None
    m = re.search(r"\(([^)]+)\)\s*$", s)
    if m:
        campus = m.group(1)
        s = s[: m.start()].strip()
    else:
        m = re.search(r"\s*-\s*[가-힣]+$", s)  # '전남대학교-광주' 형태
        if m:
            campus = m.group(0).strip(" -")
            s = s[: m.start()].strip()
    s = re.sub(r"\s*-\s*[가-힣]+$", "", s).strip()  # '한서대(태안) -항공학부' 잔여 접미
    # 분교·이원화 캠퍼스: 입결 데이터가 캠퍼스를 별도 대학으로 다루는 경우
    BRANCH = {
        ("건국대학교", "충주"): "건국대(글)",
        ("건국대학교", "글로컬"): "건국대(글)",
        ("고려대학교", "세종"): "고려대(세)",
        ("연세대학교", "원주"): "연세대(미)",
        ("한양대학교", "안산"): "한양대(에)",
        ("홍익대학교", "세종"): "홍익대(세)",
        ("상명대학교", "천안"): "상명대(천)",
        ("단국대학교", "용인"): "단국대(죽전)",
        ("단국대학교", "죽전"): "단국대(죽전)",
        ("단국대학교", None): "단국대(죽전)",
        ("단국대", None): "단국대(죽전)",
        ("단국대학교", "천안"): "단국대(천안)",
        ("동국대학교", "경주"): "동국대(W)",
        ("한국외국어대학교", "용인"): "한국외대(글)",
        ("전남대학교", "여수"): "전남대(여)",
        ("경희대학교", "용인"): "경희대",  # 국제캠은 본교 통합 모집
    }
    key = (re.sub(r"^국립", "", s), campus)
    for (base, camp), short in BRANCH.items():
        if key == (re.sub(r"^국립", "", base), camp):
            return short
    SPECIAL = {
        "한국과학기술원": "KAIST",
        "카이스트": "KAIST",
        "광주과학기술원": "GIST",
        "지스트": "GIST",
        "대구경북과학기술원": "DGIST",
        "디지스트": "DGIST",
        "울산과학기술원": "UNIST",
        "유니스트": "UNIST",
        "포항공과대학교": "POSTECH",
        "포스텍": "POSTECH",
        "한국에너지공과대학교": "한국에너지공대",
        "금오공과대학교": "금오공대",
        "서울과학기술대학교": "서울과기대",
        "한국기술교육대학교": "한국기술교대",
        "한국체육대학교": "한국체대",
        "추계예술대학교": "추계예대",
        "육군사관학교": "육사",
        "해군사관학교": "해사",
        "공군사관학교": "공사",
        "국군간호사관학교": "국간사",
        "경찰대학": "경찰대",
    }
    base = re.sub(r"^국립", "", s)
    for k, v in SPECIAL.items():
        if base == re.sub(r"^국립", "", k):
            return ("국립" + v) if s.startswith("국립") and not k.startswith("국립") and False else v
    s = s.replace("여자대학교", "여대").replace("교육대학교", "교대") \
         .replace("외국어대학교", "외대").replace("대학교", "대")
    return s


def build_resolver(index_html_path):
    """index.html의 입결 DATA에서 대학명 전체를 뽑아 '국립 유무 무시' 매칭 사전을 만든다."""
    html = open(index_html_path, encoding="utf-8").read()
    m = re.search(r'<script[^>]*id="raw-data"[^>]*>(.*?)</script>', html, re.S)
    data = json.loads(m.group(1))
    ci = data["columns"].index("대학")
    universe = set(r[ci] for r in data["rows"])
    alias = {}
    for u in universe:
        alias[u] = u
        alias.setdefault(re.sub(r"^국립", "", u), u)

    def resolve(short):
        if short is None:
            return None
        return alias.get(short) or alias.get(re.sub(r"^국립", "", short))

    return resolve


# ---------------------------------------------------------------- 전남 파일

def parse_jeonnam(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["진학결과"]
    it = ws.iter_rows(values_only=True)
    header = [clean(c) for c in next(it)]
    idx = {c: i for i, c in enumerate(header)}

    grade_cols = ["전과목", "국수영사과", "국영수사", "국영수과", "국영수", "국영사", "수영과"]
    sat_cols = ["한국사등급", "국어등급", "수학등급", "영어등급", "탐구1등급", "탐구2등급",
                "국어백분위", "수학백분위", "탐구1백분위", "탐구2백분위"]
    profile_cols = [idx[c] for c in grade_cols + sat_cols + ["국어과목", "수학과목", "탐구1과목", "탐구2과목"]]

    students, apps = [], []
    prev_profile, sid = None, 0
    for r in it:
        year = clean(r[idx["학년도"]])
        if year is None or clean(r[idx["모집시기"]]) not in ("수시", "수시1차"):
            continue
        profile = tuple(clean(r[i]) for i in profile_cols)
        if profile != prev_profile:
            sid += 1
            prev_profile = profile
            students.append({
                "id": f"J-{sid:05d}",
                "year": int(year),
                "gu": clean(r[idx["시/군"]]),  # 시 지역 / 군 지역 (읍면 지역인재 판별용)
                "grades": {c: num(r[idx[c]]) for c in grade_cols},
                "sat": {c: num(r[idx[c]]) for c in
                        ["한국사등급", "국어등급", "수학등급", "영어등급", "탐구1등급", "탐구2등급"]},
                "apps": [],
            })
        final = norm_final(r[idx["최종단계"]])
        app = {
            "univ": univ_short(r[idx["대학명"]]),
            "univ_full": clean(r[idx["대학명"]]),
            "region": clean(r[idx["지역"]]),
            "cat": norm_category(r[idx["전형유형"]]),
            "adm": clean(r[idx["전형명"]]),
            "track": norm_track(r[idx["계열"]]),
            "major": clean(r[idx["모집단위"]]),
            "stage1": norm_final(r[idx["1단계"]]),
            "final": final,
            "reg": clean(r[idx["등록여부"]]) == "등록",
        }
        students[-1]["apps"].append(len(apps))
        apps.append(app)
    wb.close()
    return {"students": students, "apps": apps}


# ---------------------------------------------------------------- 나주고 파일

def naju_headers(ws):
    """1~2행 병합 셀을 전개해 '전교과|100' 형태의 결합 헤더 리스트를 만든다."""
    ncol = ws.max_column
    h1 = [ws.cell(1, c).value for c in range(1, ncol + 1)]
    h2 = [ws.cell(2, c).value for c in range(1, ncol + 1)]
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1:
            v = ws.cell(1, mr.min_col).value
            for c in range(mr.min_col, min(mr.max_col, ncol) + 1):
                h1[c - 1] = v
    out, carry = [], ""
    for a, b in zip(h1, h2):
        a = clean(a) or ""
        b = clean(b) or ""
        # 병합이 아니라 빈 셀로 이어지는 그룹 헤더(2026 시트 수능 블록 등)도 이어받는다
        if a:
            carry = a
        elif b:
            a = carry
        out.append(f"{a}|{b}" if b and b != a else a)
    return out


def detect_offset(rows, headers):
    """헤더상 '최종단계' 위치와 실제 데이터의 합/불 값 위치를 비교해 열 밀림을 감지.

    2025 시트는 내신 점수·등급이 (일반/환산)×(점수/등급) 4열인데 헤더가 2열이라 +2 밀린다.
    """
    try:
        hpos = headers.index("최종단계")
    except ValueError:
        return 0
    from collections import Counter
    votes = Counter()
    for r in rows[:200]:
        for off in range(0, 5):
            if hpos + off < len(r) and clean(r[hpos + off]) in FINAL_OUTCOMES:
                votes[off] += 1
                break
    return votes.most_common(1)[0][0] if votes else 0


def parse_naju(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    students, apps = {}, []
    for sheet in wb.sheetnames:
        m = re.match(r"(20\d\d)", sheet)
        if not m:
            continue
        year = int(m.group(1))
        ws = wb[sheet]
        headers = naju_headers(ws)
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if clean(r[0]) in ("1", "2", "3")]
        off = detect_offset(rows, headers)

        def col(r, name, alt=None, shifted=True):
            if name not in headers and alt:
                name = alt
            if name not in headers:
                return None
            i = headers.index(name)
            if shifted and off and i >= 16:  # 밀림은 내신 점수 블록(16열)부터 발생
                i += off
            return r[i] if i < len(r) else None

        for r in rows:
            key = (year, clean(r[1]), clean(r[2]))  # 연도-반-번호 (이름은 버린다)
            if key not in students:
                students[key] = {
                    "id": f"N{year}-{len([k for k in students if k[0] == year]) + 1:03d}",
                    "year": year,
                    "grades": {
                        "전과목": num(col(r, "전교과|100")),
                        "국수영사과": num(col(r, "국영수사/과|100")),
                        "국영수사": num(col(r, "국영수사|100")),
                        "국영수과": num(col(r, "국영수과|100")),
                        "국영수": num(col(r, "국영수|100")),
                        "국영사": num(col(r, "국영사|100")),
                        "수영과": num(col(r, "수영과|100")),
                        "국어": num(col(r, "국어|100")),
                        "영어": num(col(r, "영어|100")),
                        "수학": num(col(r, "수학|100")),
                        "사회": num(col(r, "사회|100")),
                        "과학": num(col(r, "과학|100")),
                    },
                    "sat": {
                        "국어등급": num(col(r, "국어|등급")),
                        "수학등급": num(col(r, "수학|등급")),
                        "영어등급": num(col(r, "영어|등급")),
                        "탐구1등급": num(col(r, "탐구1|등급")),
                        "탐구2등급": num(col(r, "탐구2|등급")),
                        "한국사등급": num(col(r, "한국사|등급")),
                    },
                    "apps": [],
                }
            final = norm_final(col(r, "최종단계"))
            app = {
                "univ": univ_short(col(r, "대학명", shifted=False)),
                "univ_full": clean(col(r, "대학명", shifted=False)),
                "region": clean(col(r, "지역", shifted=False)),
                "cat": norm_category(col(r, "전형유형", shifted=False)),
                "adm": clean(col(r, "전형명(대)", alt="전형", shifted=False)),
                "adm_detail": clean(col(r, "세부유형", shifted=False)),
                "track": norm_track(col(r, "계열", shifted=False)),
                "major": clean(col(r, "모집단위", shifted=False)),
                "stage1": norm_final(col(r, "1단계")),
                "final": final,
                "reg": clean(col(r, "등록여부")) == "Y",
                "conv_grade": num(col(r, "내등급(환산)", alt="내등급")),  # 대학별 환산 내신
                "minreq": clean(col(r, "최저학력기준", shifted=False)),  # 수능최저 원문(2023~)
            }
            students[key]["apps"].append(len(apps))
            apps.append(app)
    wb.close()
    return {"students": list(students.values()), "apps": apps}


# ---------------------------------------------------------------- main

def main():
    jeonnam_path, naju_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    index_html = sys.argv[4] if len(sys.argv) > 4 else None
    data = {
        "jeonnam": parse_jeonnam(jeonnam_path),
        "naju": parse_naju(naju_path),
    }
    if index_html:  # 입결 데이터의 정확한 대학명으로 조인 키 확정
        resolve = build_resolver(index_html)
        n_match = 0
        for d in (data["jeonnam"], data["naju"]):
            for a in d["apps"]:
                r = resolve(a["univ"])
                if r:
                    a["univ"] = r
                    n_match += 1
        total = len(data["jeonnam"]["apps"]) + len(data["naju"]["apps"])
        print(f"입결 대학명 매칭: {n_match}/{total} ({n_match/total*100:.1f}%)")
    j, n = data["jeonnam"], data["naju"]
    print(f"전남: 학생 {len(j['students'])}명 / 지원 {len(j['apps'])}건")
    print(f"나주고: 학생 {len(n['students'])}명 / 지원 {len(n['apps'])}건")
    for label, d in (("전남", j), ("나주고", n)):
        from collections import Counter
        c = Counter(a["final"] for a in d["apps"])
        print(f"  {label} 최종단계: {dict(c)}")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"저장: {out_path}")


if __name__ == "__main__":
    main()
